from django.db.models import Count
from django.conf import settings
from django.contrib.auth.tokens import default_token_generator
from django.core.mail import send_mail
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils.encoding import force_bytes
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.safestring import mark_safe
from rest_framework import generics, status
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from rest_framework.exceptions import ParseError
from rest_framework.views import APIView
from .serializers import *
from .models import *
from openpyxl import load_workbook
import hashlib
from rest_framework.decorators import api_view, permission_classes
from django.db.models import Sum
# from .rc import getrcinput

PASSWORD_RESET_URL = (
    "https://credenz.in/forget-password/{token}/{uid}"
)

def generate_team_pass(team_id, event_name):
    password_string = f"{team_id}-{event_name}"
    hashed_string = hashlib.sha256(password_string.encode()).hexdigest()
    return hashed_string[:8]

class FeedbackView(generics.CreateAPIView):
    serializer_class = FeedbackSerializer
    queryset = Feedback.objects.all()

class UploadFileView(generics.GenericAPIView):
    permission_classes = [IsAdminUser]
    def post(self, request, format=None):
        serializer = UploadFileSerializer(data=request.data)
        if serializer.is_valid():
            file = serializer.validated_data["file"]
            wb = load_workbook(filename=file)
            ws = wb.active
            upi_col_num = None
            for cell in ws['1']:
                if cell.value == 'upi':
                    upi_col_num = cell.column
                    break
            f_check = 0
            c_check = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                upi = str(row[upi_col_num - 1])
                if upi != "None":
                    transaction_id = upi.split('/')[1].strip()
                    amount_excel = row[upi_col_num - 2]
                    print(transaction_id, amount_excel)
                try:
                    transaction = Transaction.objects.get(transaction_id=transaction_id)
                except Transaction.DoesNotExist:
                    continue
                if transaction.payment == "CO":
                    continue
                amount_transaction = transaction.amount
                if amount_excel == amount_transaction:
                    c_check += 1
                    Order.objects.filter(transaction_id=transaction_id).update(payment="CO")
                    Transaction.objects.filter(transaction_id=transaction_id).update(payment="CO")

                    # coin update for referred user
                    if Referral.objects.filter(referred_user = transaction.user):
                        referral_obj = Referral.objects.get(referred_user = transaction.user)
                        coin_user = User.objects.get(username = referral_obj.referrer.username)
                        coin_user.coins += int(transaction.amount / 10)
                        coin_user.save()
                        referral_obj.save()
                    
                    # send confirmation email
                    user = transaction.user
                    email = user.email
                    event_l = transaction.event_list
                    events = Event.objects.filter(event_id__in=event_l)
                    context = {"user": user, "events": events, "transaction_id":transaction_id, "order_date": transaction.order_date}
                    html_message = render_to_string("confirmation.html", context=context)
                    try:
                        send_mail(
                        'Order Confirmation',
                        '',
                        settings.EMAIL_HOST_USER,
                        [email],
                        html_message=html_message,
                        fail_silently=False,
                        )
                    except Exception as e:
                        print(f"Email failed due to: {e}")
                else:
                    f_check += 1
            return Response({"message" : f"transactions completed: faulty {f_check}, completed {c_check}"}, status=status.HTTP_202_ACCEPTED)

@api_view(['GET'])
def verifyuser(request, uidb64, token):
    try:
        uid = force_bytes(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is not None and default_token_generator.check_token(user, token):
        user.is_active = True
        user.save()
        refresh = RefreshToken.for_user(user)
        token_pair = {
            'refresh': str(refresh),
            'access': str(refresh.access_token),
        }
        return Response({'message': 'Email verified successfully.','token':token_pair['access']}, status=status.HTTP_200_OK)
    else:
        return Response({'message': 'Invalid verification link.'}, status=status.HTTP_400_BAD_REQUEST)

class UserRegisterView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer

    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data)
        if (serializer.is_valid()):
            #serializer.is_valid(raise_exception=True)
            user = serializer.save()
            refresh = RefreshToken.for_user(user)
            token_pair = {
                'refresh': str(refresh),
                'access': str(refresh.access_token),
            }
            return Response(token_pair, status=status.HTTP_201_CREATED)
            # return Response({'message':'Verification mail has been sent.'}, status=status.HTTP_201_CREATED)
        else:
            return Response({'message':serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

class UserLoginView(generics.GenericAPIView):
    serializer_class = TokenPairSerializer

    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)
        return Response(serializer.validated_data, status=status.HTTP_200_OK)
    
class EventsDetail(generics.ListAPIView):
    queryset = Event.objects.all()
    serializer_class = EventSerializer

class ProfileView(generics.RetrieveAPIView):
    serializer_class = ProfileSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        return self.request.user
    
class LeaderboardView(generics.ListAPIView):
    queryset = (
        User.objects.all().order_by("-coins")[:10]
    )

    serializer_class = ProfileSerializer

class PasswordResetView(generics.GenericAPIView):
    def post(self, request):
        email = request.data.get("email")

        try:
            user = User.objects.get(email=email)
            uid = urlsafe_base64_encode(force_bytes(user.id))
            print(uid)
            token = default_token_generator.make_token(user)
            print(token)
            reset_url = mark_safe(PASSWORD_RESET_URL.format(token=token, uid=uid))
            context = {"user": user, "reset_url": reset_url}
            html_message = render_to_string("password-reset.html", context=context)
            send_mail(
                'Reset your password',
                '',
                settings.EMAIL_HOST_USER,
                [email],
                html_message=html_message,
                fail_silently=False,)

            return Response(
                {"message": "Link sent to mail!"}, status=status.HTTP_200_OK
            )

        except User.DoesNotExist:
            raise ValidationError({"email": "This email dosn't belongs to any user"})


class PasswordResetConfirmView(generics.GenericAPIView):
    def post(self, request):
        uid = request.data["uid"]
        user_id = urlsafe_base64_decode(uid)
        token = request.data["token"]
        new_password = request.data["new_password"]
        user = User.objects.get(id=user_id)
        if not user:
            message = "User doesn't exist"
            return Response({"message": message}, status=status.HTTP_400_BAD_REQUEST)
        elif default_token_generator.check_token(user, token):
            user.set_password(new_password)
            user.save()
            return Response(
                {"message": "password reset success"}, status=status.HTTP_200_OK
            )
        else:
            message = "Link is invalid or expired"
        return Response({"message": message}, status=status.HTTP_400_BAD_REQUEST)
    
class VerifyReferralCodeView(generics.GenericAPIView):
    serializer_class = ReferralCodeVerifySerializer

    def post(self, request):
        serializer = self.get_serializer(data=self.request.data)
        serializer.is_valid(raise_exception=True)
        return Response(serializer.validated_data)

class OrderView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        user = get_object_or_404(User, id=request.user.id)
        orders = Order.objects.filter(user=user)
        serializer = OrderSerializer(orders, many=True)
        return Response(serializer.data)

class PlaceOrderView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        # save order for each particular event
        user = request.user
        event_list = request.data['event_list']
        transaction_id = request.data["transaction_id"]
        amount = request.data["amount"]
        
        if Transaction.objects.filter(transaction_id=transaction_id).exists():
            return Response({"message" : "Invalid Transaction Id."}, status=status.HTTP_400_BAD_REQUEST)

        if Transaction.objects.filter(transaction_id=transaction_id).exists():
            return Response({"message" : "Invalid Transaction Id."}, status=status.HTTP_400_BAD_REQUEST)

        order_list = []
        for event in event_list:
            if not Order.objects.filter(user = request.user, event = Event.objects.get(event_id = event)).first():
                order = Order(user = user, event = Event.objects.get(event_id = event), transaction_id=transaction_id)
                order_list.append(order)
            else:
                return Response({"message" : "order already placed"})
    
            # add transaction for the complete list of events
        if not Transaction.objects.filter(transaction_id = transaction_id).first():
            transaction = Transaction(event_list = event_list, user = user, transaction_id = transaction_id, amount=amount)
            transaction.save()

            # check if every event orderis valid and then save
            Order.objects.bulk_create(order_list)

            # send order email
            events = Event.objects.filter(event_id__in=event_list)
            user = request.user
            email = user.email
            context = {"user": user, "events": events, "transaction_id":transaction_id}
            html_message = render_to_string("event.html", context=context)
            try:
                send_mail(
                'Your Order',
                '',
                settings.EMAIL_HOST_USER,
                [email],
                html_message=html_message,
                fail_silently=False,
                )
            except Exception as e:
                print(f"Email failed due to: {e}")
        else:
            transaction = Transaction(event_list = event_list, user = user, transaction_id = transaction_id, amount=amount, payment="IO")
            transaction.save()
            return Response({"message" : "Transaction already performed!"})

        return Response({"message" : "order placed"}, status=status.HTTP_201_CREATED)
    

    
# Team event APIs
class GenerateTeamCodeView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        event = int(request.data['event_id'])
        team_name = request.data['team_name']
        user = request.user

        if Order.objects.filter(event=event, user=user):
            if not user.team_set.filter(event=event).first():
                new_team = Team.objects.create(event=Event.objects.get(event_id = event), team_name = team_name)
                new_team.user.add(request.user)
                team_password = generate_team_pass(new_team.team_id, Event.objects.filter(event_id=event))
                new_team.team_password = team_password
                new_team.save()

                # send team details
                user = request.user
                email = user.email
                event_check = Event.objects.filter(event_id=event)
                context = {"user": user, "team_id": new_team.team_id, "event": event_check, "team_password" : new_team.team_password}
                html_message = render_to_string("team.html", context=context)
                try:
                    send_mail(
                            'Your Team',
                            '',
                            settings.EMAIL_HOST_USER,
                            [email],
                            html_message=html_message,
                            fail_silently=False,
                        )
                except Exception as e:
                    print(f"email failed due to: {e}")
                return Response({"message": new_team.team_id}, status=status.HTTP_201_CREATED)
            else:
                return Response({"message" : "Team already exists for this user and event."})
        else:
            return Response({"message" : "No order exists for this user and event."})

class JoinTeamView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        team_id = request.data['team_id']
        event_check = Team.objects.filter(team_id=team_id).first()

        if Team.objects.filter(team_id = team_id):
            event_c = Team.objects.filter(team_id=team_id).first()

            if event_c.event.event_id == "101" or event_c.event.event_id == "102":
                if Order.objects.filter(event = event_check.event, user = user):
                    if not Team.objects.filter(user = user, event = event_check.event):
                            exis_team = Team.objects.get(team_id = team_id)
                            if exis_team.user.count() >= exis_team.event.group_size:
                                return Response({"message" : "Maximum team size reached already!"}, status=status.HTTP_400_BAD_REQUEST)
                            else:
                                user_one = exis_team.user.first()
                                if user_one.senior != request.user.senior:
                                    return Response({"message" : "Users must be from the same category to form a team"}, status=status.HTTP_400_BAD_REQUEST)
                                exis_team.user.add(request.user)

                                # send team email
                                user = request.user
                                email = user.email
                                context = {"user": user, "team_id": team_id, "event": event_check, "team_password" : exis_team.team_password}
                                html_message = render_to_string("team.html", context=context)
                                try:
                                    send_mail(
                                    'Your Team',
                                    '',
                                    settings.EMAIL_HOST_USER,
                                    [email],
                                    html_message=html_message,
                                    fail_silently=False,
                                    )
                                except Exception as e:
                                    print(f"Email failed due to: {e}")
                                return Response({"message" : "User added to Team"}, status=status.HTTP_201_CREATED)

                    else:
                        return  Response({"message" : "Team already exists for this user and event."})
                else:
                    return Response({"message" : "No order exists for this user and event."})
            else:
                if not Team.objects.filter(user = user, event = event_check.event):
                    exis_team = Team.objects.get(team_id = team_id)
                    if exis_team.user.count() >= exis_team.event.group_size:
                        return Response({"message" : "Maximum team size reached already!"}, status=status.HTTP_400_BAD_REQUEST)
                    else:
                        exis_team.user.add(request.user)

                        # send team email
                        user = request.user
                        email = user.email
                        context = {"user": user, "team_id": team_id, "event": event_check, "team_password" : exis_team.team_password}
                        html_message = render_to_string("team.html", context=context)
                        try:
                            send_mail(
                            'Your Team',
                            '',
                            settings.EMAIL_HOST_USER,
                            [email],
                            html_message=html_message,
                            fail_silently=False,
                            )
                        except Exception as e:
                            print(f"Email failed due to: {e}")
                        return Response({"message" : "User added to Team"}, status=status.HTTP_201_CREATED)
                else:
                    return  Response({"message" : "Team already exists for this user and event."})
        else:
            return Response({"message" : "Invalid team ID"})

class TeamView(generics.ListAPIView):
    serializer_class = TeamSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        return Team.objects.filter(user=user)
    
# Offline Register APIs
class RegisterPlayerView(generics.CreateAPIView):
    permission_classes = [IsAuthenticated]

    queryset = User.objects.all()
    serializer_class = UserSerializer

    def post(self, request, *args, **kwargs):
        off_officer = request.user
        
        if ( User.objects.get(username = off_officer.username).offline_officer== True):
            serializer = self.serializer_class(data=request.data)
            serializer.is_valid(raise_exception=True)
            user = serializer.save()

            # Sending email with user details
            email = user.email
            context = {"user": user, "username": user.username, "password": request.data.get("password")} 
            html_message = render_to_string("offline-register.html", context=context)
            send_mail(
                'Your Registration Details',
                '',
                settings.EMAIL_HOST_USER,
                [email],
                html_message=html_message,
                fail_silently=False,
            )

            return Response({"message": "User saved and email sent successfully"}, status=status.HTTP_201_CREATED)        
        return Response({"message" : "Not an offline officer"}, status=status.HTTP_203_NON_AUTHORITATIVE_INFORMATION)
        
class OfflineOrderView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        order_user = request.data['username']
        event_list = request.data['event_list']
        transaction_id = request.data["transaction_id"]
        amount = request.data["amount"]
        off_officer = request.user
        event_list = list(event_list)
        
        if Transaction.objects.filter(transaction_id=transaction_id).exists():
            return Response({"message" : "Invalid Transaction Id."}, status=status.HTTP_400_BAD_REQUEST)

        if (User.objects.get(username = off_officer.username).offline_officer== True):
            
            for event in event_list:
                if not Order.objects.filter(user = User.objects.get(username=order_user), event = Event.objects.get(event_id = event)).first():
                    order = Order(user = User.objects.get(username=order_user), event = Event.objects.get(event_id = event), order_taker = request.user.full_name, payment="CO", transaction_id=transaction_id)
                    order.save()

            # add transaction for the complete list of events
            if not Transaction.objects.filter(transaction_id=transaction_id).first():
                transaction = Transaction(event_list = event_list, user = User.objects.get(username=order_user), transaction_id = transaction_id, amount=amount, payment ="CO")
                transaction.save()

             # send order email
            events = Event.objects.filter(event_id__in=event_list)
            user = User.objects.filter(username=order_user).first()
            email = user.email
            context = {"user": user, "events": events, "transaction_id":transaction_id}
            html_message = render_to_string("event.html", context=context)
            try:
                send_mail(
                'Your Order',
                '',
                settings.EMAIL_HOST_USER,
                [email],
                html_message=html_message,
                fail_silently=False,
                )
            except Exception as e:
                print(f"Email failed due to: {e}")

            return Response({"message" : "order placed"}, status=status.HTTP_201_CREATED)
        
        return Response({"message" : "Transaction not allowed"})

class TransactionListView(generics.ListAPIView):
    serializer_class = TransactionSerializer
    permission_classes = [IsAdminUser]
    queryset = Transaction.objects.all()
    

class TransactionConfirmView(generics.GenericAPIView):
    permission_classes = [IsAdminUser]

    def post(self, request):
        transaction_id = request.data["transaction_id"]

        Transaction.objects.filter(transaction_id = transaction_id).update(payment="CO")
        Order.objects.filter(transaction_id=transaction_id).update(payment="CO")

        transaction = Transaction.objects.get(transaction_id=transaction_id)

        # coin update for referred user
        if Referral.objects.filter(referred_user = transaction.user):
            referral_obj = Referral.objects.get(referred_user = transaction.user)
            coin_user = User.objects.get(username = referral_obj.referrer.username)
            coin_user.coins += int(transaction.amount / 10)
            coin_user.save()
            referral_obj.save()

        # send confirmation email
        user = transaction.user
        email = user.email
        event_l = transaction.event_list
        events = Event.objects.filter(event_id__in=event_l)
        context = {"user": user, "events": events, "transaction_id":transaction_id, "order_date": transaction.order_date}
        html_message = render_to_string("confirmation.html", context=context)
        try:
            send_mail(
                'Order Confirmation',
                '',
                settings.EMAIL_HOST_USER,
                [email],
                html_message=html_message,
                fail_silently=False,
            )
        except Exception as e:
            # Handle the exception here
            print(f"An error occurred while sending the email: {e}")
               

        return Response({"message" : "confirmed"}, status=status.HTTP_202_ACCEPTED)
    
class AdminPassView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        order_user = request.data['username']
        transaction_id = request.data["transaction_id"]
        amount = request.data["amount"]
        off_officer = request.user
        
        if Transaction.objects.filter(transaction_id=transaction_id).exists():
            return Response({"message" : "Invalid Transaction Id."}, status=status.HTTP_400_BAD_REQUEST)

        if (User.objects.get(username = off_officer.username).offline_officer== True):
            event_id_list = Event.objects.values_list('event_id', flat=True)
            event_list = [ int(event_id)  for event_id in event_id_list if int(event_id) !=111 ]
            for event in event_list:
                if not Order.objects.filter(user = User.objects.get(username=order_user), event = Event.objects.get(event_id = event)):
                    order = Order(user = User.objects.get(username=order_user), event = Event.objects.get(event_id = event), order_taker = request.user.full_name, payment="CO", transaction_id=transaction_id)
                    order.save()

            # add transaction for the complete list of events
            Order.objects.filter(user=User.objects.get(username=order_user)).update(payment="CO")
            transaction = Transaction(event_list = event_list, user = User.objects.get(username=order_user), transaction_id = transaction_id, amount=amount, payment ="CO")
            transaction.save()

             # send order email
            events = Event.objects.filter(event_id__in=event_list)
            user = User.objects.filter(username=order_user).first()
            email = user.email
            user.has_purchased_pass = True
            user.save()
            context = {"user": user, "events": events, "transaction_id":transaction_id}
            html_message = render_to_string("event.html", context=context)
            try:
                send_mail(
                'Your Order',
                '',
                settings.EMAIL_HOST_USER,
                [email],
                html_message=html_message,
                fail_silently=False,
                )
            except Exception as e:
                print(f"Email failed due to: {e}")

            return Response({"message" : "order placed"}, status=status.HTTP_201_CREATED)
        
        return Response({"message" : "Transaction not allowed"})

class PassView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):

        transaction_id = request.data["transaction_id"]
        amount = request.data["amount"]
        buyer = request.user
        event_id_list = Event.objects.values_list('event_id', flat=True)
        event_list = [ int(event_id)  for event_id in event_id_list if int(event_id) !=111 ]

        if Transaction.objects.filter(transaction_id=transaction_id).exists():
            return Response({"message" : "Invalid Transaction Id."}, status=status.HTTP_400_BAD_REQUEST)

        for event in event_list:
            if not Order.objects.filter(user = buyer, event = Event.objects.get(event_id = event)):
                order = Order(user = buyer, event = Event.objects.get(event_id = event), transaction_id=transaction_id)
                order.save()

        # add transaction for the complete list of events
        transaction = Transaction(event_list = event_list, user = buyer, transaction_id = transaction_id, amount=amount)
        transaction.save()

        # send order email
        events = Event.objects.filter(event_id__in=event_list)
        email = buyer.email
        buyer.has_purchased_pass = True
        buyer.save()
        context = {"user": buyer, "events": events, "transaction_id":transaction_id}
        html_message = render_to_string("event.html", context=context)
        try:
            send_mail(
                'Your Order',
                '',
                settings.EMAIL_HOST_USER,
                [email],
                html_message=html_message,
                fail_silently=False,
            )
        except Exception as e:
            print(f"Email failed due to: {e}")

        return Response({"message" : "order placed"}, status=status.HTTP_201_CREATED)
        
class ValidateUserView(APIView):

    def post(self, request):
        try:
            username = request.data["username"]
            password = request.data["password"]
            event = request.data["event"]
            is_team = bool(request.data["is_team"])
            print(is_team)
            if event not in ["clash", "rc", "wallstreet"]:
                raise ParseError("Invalid event")
        except KeyError:
            raise ParseError("Username, password, event and is_team fields are required.")

        if is_team:
            team = Team.objects.filter(team_id=username).first()
            if team:
                if team.team_password != password:
                    return Response({"detail": "incorrect password"}, status=status.HTTP_401_UNAUTHORIZED)
                users = team.user.all()
                serializer = UserSerializer(users, many=True)
                return Response({"users" :serializer.data, "detail":"verified"}, status=status.HTTP_200_OK)
            else:
                return Response({"detail": "team does not exist"}, status=status.HTTP_404_NOT_FOUND)
        
        user = User.objects.filter(username=username).first()
        
        if user:
            order = Order.objects.filter(user=user, event=Event.objects.get(event_name=event)).first()
            if not order:
                return Response({"detail": f"user is not registered for {event}"}, status=status.HTTP_401_UNAUTHORIZED)
            if user.check_password(password):
                serializer = UserSerializer(user)
                return Response({"user" : serializer.data, "detail" : "verified"}, status=status.HTTP_200_OK)
            else:
                return Response({"detail": "incorrect password"}, status=status.HTTP_401_UNAUTHORIZED)
        else:
            return Response({"detail": "user does not exist"}, status=status.HTTP_404_NOT_FOUND)
        


@api_view(['GET'])
@permission_classes([IsAdminUser])
def total_amount(request):
    total = Transaction.objects.aggregate(Sum('amount'))['amount__sum']
    return Response({'total_amount': total})

# class RcAPI(generics.GenericAPIView):
#     def post(self, request):
#         rc_input = request.data['rc_input']
#         rc_ques = request.data['rc_ques']

#         rc_output = getrcinput(rc_ques, rc_input)

#         return Response({"rc_output" : rc_output}, status=status.HTTP_200_OK)

